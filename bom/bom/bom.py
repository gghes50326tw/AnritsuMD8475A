import os
import re
import csv
import sys
import time
import openpyxl
import warnings
import configparser
import pandas as pd


class getoutofloop(Exception):
    pass


warnings.filterwarnings('ignore')
ver = '1.0.2'
dd = '2022-07-05'
info_format = "{:>14s}{:<20s}"
print(info_format.format('Last updated: ', dd))
print(info_format.format('Version: ', ver))
print('')

# 20220427 initial version
config = configparser.ConfigParser()
config.read('config.ini')
file_supplier = config['Filename']['supplier']
file_factory = config['Filename']['factory']
file_xdm = config['Filename']['xdm']

# Check file_supplier and file_xdm in the working directory
local_dir = os.path.dirname(os.path.realpath(__file__))
if not os.path.exists(file_supplier):
    print('Can\'t find source file : %s.' % file_supplier)
    print('Please check the file exists in %s.' % local_dir)
    sys.exit(0)
elif not os.path.exists(file_xdm):
    print('Can\'t find destination file : %s.' % file_xdm)
    print('Please check the file exists in %s.' % local_dir)
    sys.exit(0)

# Define fields
field_dest = {'Tier 1 Supplier PN': '',
              'Tier 1 Supplier Part Description': '',
              'QPA': '',
              'Sub-tier Supplier ID': '',
              'Sub-tier Supplier Name': '',
              'Sub-tier Supplier PN': '',
              'Factory Name': '',
              'Factory Address Line 1': '',
              'Factory City': '',
              'Factory State/Province': '',
              'Factory Postcode/Zipcode': '',
              'Factory Country': ''}

field_supplier = {'Material': 'Tier 1 Supplier PN',
                  'Descripition': 'Tier 1 Supplier Part Description',
                  'BOM Qty': 'QPA',
                  'Vendor 1': 'Sub-tier Supplier ID',
                  'Manufacturer of Vendor1': 'Sub-tier Supplier Name',
                  'Manuf Part No of Vendor 1': 'Sub-tier Supplier PN'}

field_factory = {'fac_id': 'Sub-tier Supplier ID',
                 'fac_pn': 'Sub-tier Supplier PN',
                 'fac_name': 'Factory Name',
                 'fac_addr': 'Factory Address Line 1',
                 'fac_city': 'Factory City',
                 'fac_state': 'Factory State/Province',
                 'fac_zipcode': 'Factory Postcode/Zipcode',
                 'fac_country': 'Factory Country'}

# Get header row no
wb_supplier = openpyxl.load_workbook(file_supplier, read_only=True)
ws_supplier = wb_supplier.active
index_no = 0
try:
    for row in ws_supplier.iter_rows():
        for cell in row:
            if cell.value == 'Material':
                raise getoutofloop
            else:
                continue
        index_no += 1
except getoutofloop:
    pass

# Retrieve necessary columns from supplier file
cols = []
cnt = 0
for i in ws_supplier[index_no+1]:
    if i.value in field_supplier:
        cols.append(cnt)
    cnt += 1
df_sup = pd.read_excel(file_supplier, header=index_no, usecols=cols)

# Check point: Mandatory fields are existed in file_supplier.
md_col = ['Material', 'Descripition', 'BOM Qty', 'Vendor 1',
          'Manufacturer of Vendor1', 'Manuf Part No of Vendor 1']
src_col = []
for item in ws_supplier[index_no+1]:
    src_col.append(item.value)
for field in md_col:
    if field in src_col:
        pass
    else:
        print('%s does not include mandatory column: %s' % (file_supplier, field))
        sys.exit(0)
wb_supplier._archive.close()

# Data validation: file_factory
df_fac = pd.read_excel(file_factory)
fac_cols = df_fac.columns
fac_data = df_fac.to_dict('records')
dup_label = 'Sub-tier Supplier ID'
df_fac.duplicated(subset=dup_label, keep=False)
new_fac = df_fac.drop_duplicates(subset=dup_label, keep=False).to_dict('records')
index_fac = {}
for records in new_fac:
    index_fac[records[dup_label]] = [records['Factory Name'],
                                     records['Factory Address Line 1'],
                                     records['Factory City'],
                                     records['Factory State/Province'],
                                     records['Factory Postcode/Zipcode'],
                                     records['Factory Country']]

rawdata = df_sup.to_dict('records')

# get column index from dest file
df_dst = pd.read_excel(file_xdm, sheet_name='Input')

for name in field_dest:
    try:
        field_dest[name] = int(df_dst.columns.get_loc(name)) + 1
    except Exception:
        pass

myworkbook = openpyxl.load_workbook(file_xdm)
ws = myworkbook.get_sheet_by_name('Input')

# write cell
start_row = 2
pattern = r'.+(-.+)'

for i in rawdata:
    if i['Vendor 1'] in index_fac:
        tmp = {field_factory['fac_name']: index_fac[i['Vendor 1']][0],
               field_factory['fac_addr']: index_fac[i['Vendor 1']][1],
               field_factory['fac_city']: index_fac[i['Vendor 1']][2],
               field_factory['fac_state']: index_fac[i['Vendor 1']][3],
               field_factory['fac_zipcode']: index_fac[i['Vendor 1']][4],
               field_factory['fac_country']: index_fac[i['Vendor 1']][5]}
        i.update(tmp)

for data in rawdata:
    index = 0
    write_row = list(data.values())
    for col in field_dest.keys():
        try:
            if col == 'Sub-tier Supplier ID':
                if re.fullmatch(pattern, write_row[index]):
                    matchstr = re.findall(pattern, write_row[index])
                    write_row[index] = 'LCM1%s' % (matchstr[0])
                elif write_row[index] == '':
                    pass
                else:
                    write_row[index] = 'LCM1-%s' % write_row[index]
            ws.cell(row=start_row, column=field_dest[col],
                    value=write_row[index])
            index += 1
            continue
        except Exception as ex:
            pass
    start_row += 1

# save
timestr = time.strftime("%Y%m%d_%H%M%S", time.localtime())
file_xdm = '%s_%s' % (timestr, file_xdm)
myworkbook.save(file_xdm)
print('File is save to: %s/%s' % (local_dir, file_xdm))
