# -*- coding: utf-8 -*-
"""
Created on Wed Nov 30 09:30:35 2022

@author: Vivian
"""
import openpyxl
import pandas as pd
field_dest = {'Tier 1 Supplier PN': '',
              'Tier 1 Supplier Part Description': '',
              'QPA': '',
              'Sub-tier Supplier ID': '',
              'Sub-tier Supplier Name': '',
              'Sub-tier Supplier PN': '',
              'Factory Name': '',
              'Factory Address Line 1': '',
              'Factory City': '',
              'Factory State/Province': '',
              'Factory Postcode/Zipcode': '',
              'Factory Country': ''}

field_naterial = {'Material': 'Tier 1 Supplier PN',
                  'Descripition': 'Tier 1 Supplier Part Description',
                  'BOM Qty': 'QPA',
                  'Vendor 1': 'Sub-tier Supplier ID',
                  'Manufacturer of Vendor1': 'Sub-tier Supplier Name',
                  'Manuf Part No of Vendor 1': 'Sub-tier Supplier PN'}

field_factory = {'fac_id': 'Sub-tier Supplier ID',
                 'fac_pn': 'Sub-tier Supplier PN',
                 'fac_name': 'Factory Name',
                 'fac_addr': 'Factory Address Line 1',
                 'fac_city': 'Factory City',
                 'fac_state': 'Factory State/Province',
                 'fac_zipcode': 'Factory Postcode/Zipcode',
                 'fac_country': 'Factory Country'}

xdm = openpyxl.load_workbook('xdm1.xlsx')
s1 = xdm['Input']

df_fac = pd.read_excel("Demo_factory.xlsx")
df_mat = pd.read_excel("Demo_Material.xlsx")

fac_listkey=[field_factory['fac_id'],field_factory['fac_pn'],field_factory['fac_name'],field_factory['fac_addr'],
             field_factory['fac_city'],field_factory['fac_state'],field_factory['fac_zipcode'],field_factory['fac_country']]
fac_listvalue=[]
fac={}

for i in range(df_fac.shape[0]):
    #if df_fac.iloc[i,1]==df_mat.iloc[i,6]:
        for j in range(df_fac.shape[1]):
            fac_listvalue.append(str(df_fac.iloc[i,j]))
            fac[fac_listkey[j]]=fac_listvalue[j]
            for k in range(s1.max_column):    
                    if s1.cell(1,k+1).value==fac_listkey[j]:
                        s1.cell(i+2,k+1).value = fac[fac_listkey[j]]
                        print(fac[fac_listkey[j]])
            if len(fac_listvalue)==df_fac.shape[1]:
                fac_listvalue.clear()
xdm.save('xdm1.xlsx')
